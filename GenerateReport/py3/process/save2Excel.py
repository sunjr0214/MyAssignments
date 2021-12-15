#!/usr/bin/env python
# -*- coding:utf-8 -*-
from __future__ import division

import os
import sys
import time

import pandas as pd
from openpyxl import load_workbook

currentPath = os.path.dirname(os.path.realpath(__file__))
py3Path = os.path.abspath(os.path.join(currentPath, '..'))
sys.path.append(py3Path)
from py3.constant import commonConst as coco
from py3.utils import Path as pth


# if sys.getdefaultencoding() != 'utf-8':
#     reload(sys)
#     sys.setdefaultencoding('utf-8')


def saveOneCitySingleMonthData(TemplateExcPrefix, saveExcPrefix, riskRankBigTab, selectedCols, firstCellLoc):
    # 选取字段
    constObj = coco.colNameConst()
    numberMap = constObj.numberMap

    # 生成Excel
    nowDate = time.strftime("%Y-%m-%d-%Hh%Mm%Ss", time.localtime())
    TemplateExcName = '{}.xlsx'.format(TemplateExcPrefix)
    saveExcName = '{}_{}.xlsx'.format(saveExcPrefix, nowDate)
    templatePath = pth.getTemplateFilePath(TemplateExcName)
    savePath = pth.getOneCitySingleMonthReportSavePath(saveExcName)

    write2excel(templatePath, savePath, riskRankBigTab, numberMap, selectedCols, firstCellLoc)
    return savePath


def saveSingleMonthData(TemplateExcPrefix, saveExcPrefix, riskRankBigTab, selectedCols, firstCellLoc, theme):
    # 选取字段
    constObj = coco.colNameConst()
    numberMap = constObj.numberMap

    # 生成Excel
    nowDate = time.strftime("%Y-%m-%d-%Hh%Mm%Ss", time.localtime())
    TemplateExcName = '{}.xlsx'.format(TemplateExcPrefix)
    saveExcName = '{}_{}.xlsx'.format(saveExcPrefix, nowDate)
    templatePath = pth.getTemplateFilePath(TemplateExcName)
    savePath = pth.getSingleMonthReportSavePath(saveExcName)

    write2excel(templatePath, savePath, riskRankBigTab, numberMap, selectedCols, firstCellLoc, theme)
    return savePath


def saveMultiMonthData(TemplateExcPrefix, saveExcPrefix, riskRankBigTab, selectedCols, firstCellLoc, theme):
    # 选取字段
    constObj = coco.colNameConst()
    numberMap = constObj.numberMap

    # 生成Excel
    nowDate = time.strftime("%Y-%m-%d-%Hh%Mm%Ss", time.localtime())
    TemplateExcName = '{}.xlsx'.format(TemplateExcPrefix)
    saveExcName = '{}_{}.xlsx'.format(saveExcPrefix, nowDate)
    templatePath = pth.getTemplateFilePath(TemplateExcName)
    savePath = pth.getMultiMonthReportSavePath(saveExcName)

    write2excel(templatePath, savePath, riskRankBigTab, numberMap, selectedCols, firstCellLoc, theme)
    return savePath


# 写入excel模板
def write2excel(templatePath, savePath, DF, numberMap, selectedCols, firstCellLoc, *theme):
    wb = load_workbook(templatePath)
    sheets = wb.sheetnames
    workSheet = wb[sheets[0]]  # 默认选取第一个sheet
    rowNum = DF.shape[0]
    colNum = len(selectedCols)

    # 写入序号列
    for i in range(0, rowNum):
        workSheet.cell(row=i + firstCellLoc[0], column=firstCellLoc[1]).value = i+1

    # 写入非序号列
    for i in range(0, colNum):
        colCode = selectedCols[i]
        colName = numberMap[colCode]
        currentDF = pd.DataFrame(DF[colName])
        for j in range(0, rowNum):
            singleValue = currentDF.iloc[j, 0]  # 选取指定列遍历插入
            workSheet.cell(row=j + firstCellLoc[0], column=i + firstCellLoc[1] + 1).value = singleValue

    if len(theme) != 0:
        # 写入主题行
        workSheet.cell(row=1, column=2).value = theme[0]

    wb.save(savePath)


def backUpNetCheck(DF, bakupNetCheckExcName):
    bakupNetCheckPath = pth.getPostNetCheckPath(bakupNetCheckExcName)
    try:
        DF.to_excel(bakupNetCheckPath)
    except Exception:
        raise Exception('保存失败')
